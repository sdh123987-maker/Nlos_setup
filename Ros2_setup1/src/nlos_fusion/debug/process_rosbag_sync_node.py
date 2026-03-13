#!/usr/bin/env python3
"""
==============================================
NLOS Multi-Sensor Rosbag Processor v2
==============================================
v5 (멀티레이더 + Hold) 구조에 맞춤

토픽 구조:
  /sync/radar/merged/points     - 통합 포인트클라우드
  /sync/radar_0/points          - 레이더0 개별
  /sync/radar_0/heatmap         - 레이더0 히트맵
  /sync/radar_1/points          - 레이더1 개별
  /sync/radar_1/heatmap         - 레이더1 히트맵
  /sync/camera/color
  /sync/audio/spectrum
"""

import os
import sys
import glob
import cv2
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
import traceback
import warnings
import re
warnings.filterwarnings('ignore')

try:
    from rclpy.serialization import deserialize_message
    from rosbag2_py import SequentialReader, StorageOptions, ConverterOptions
    from sensor_msgs.msg import PointCloud2, Image, Imu
    from std_msgs.msg import Float32MultiArray, Float32, String
    from sensor_msgs_py import point_cloud2
    HAS_ROS = True
except ImportError:
    HAS_ROS = False
    print("⚠️ ROS2 라이브러리가 필요합니다.")

# ==========================================
# 설정
# ==========================================
WORKSPACE_DIR = os.path.expanduser('~/Nlos/Ros2_setup')
ROSBAG_DIR = os.path.join(WORKSPACE_DIR, 'rosbag')
OUTPUT_BASE_DIR = os.path.join(WORKSPACE_DIR, 'processed_data')

MAX_RANGE_METERS = 10.0
VIRTUAL_ANTENNAS = 12
VELOCITY_THRESHOLD = 0.1

AUDIO_ANGLE_MIN, AUDIO_ANGLE_MAX = 0, 180


def get_latest_bag_path(base_dir):
    all_bags = glob.glob(os.path.join(base_dir, 'sync_*'))
    if not all_bags:
        print("❌ Rosbag 폴더가 비어있습니다.")
        return None
    return max(all_bags, key=os.path.getctime)


class BagProcessor:
    def __init__(self, bag_path):
        self.bag_path = bag_path
        self.bag_name = os.path.basename(bag_path)
        self.output_dir = os.path.join(OUTPUT_BASE_DIR, self.bag_name)
        
        # 기본 폴더
        self.dirs = {
            'radar': os.path.join(self.output_dir, 'radar'),
            'camera': os.path.join(self.output_dir, 'camera'),
            'camera_images': os.path.join(self.output_dir, 'camera', 'images'),
            'acoustic': os.path.join(self.output_dir, 'acoustic'),
            'acoustic_csv': os.path.join(self.output_dir, 'acoustic', 'csv'),
            'acoustic_plots': os.path.join(self.output_dir, 'acoustic', 'plots'),
        }
        
        # 레이더별 폴더 (동적 생성)
        self.radar_dirs = {}
        
        for d in self.dirs.values():
            os.makedirs(d, exist_ok=True)

        # 데이터 저장소
        self.data_store = {
            'frame_info': []
        }
        
        # 레이더별 데이터
        self.radar_data = {}  # radar_id -> {'points': [], 'range': [], 'noise': []}
        
        # 순차 프레임 카운터
        self.seq_counts = {'merged': 0, 'camera': 0, 'audio': 0}
        self.radar_seq_counts = {}  # radar_id -> count
        
        # 이미지 생성 오류 추적
        self.image_errors = {}

    def setup_radar_dirs(self, radar_id):
        """레이더별 폴더 동적 생성"""
        if radar_id not in self.radar_dirs:
            base = os.path.join(self.dirs['radar'], radar_id)
            self.radar_dirs[radar_id] = {
                'csv': os.path.join(base, 'csv'),
                'pcd': os.path.join(base, 'pcd'),
                'img': os.path.join(base, 'img'),
                'heatmap': os.path.join(base, 'heatmap'),
            }
            for d in self.radar_dirs[radar_id].values():
                os.makedirs(d, exist_ok=True)
            
            self.radar_data[radar_id] = {'points': [], 'range': [], 'noise': []}
            self.radar_seq_counts[radar_id] = 0
            self.image_errors[f'{radar_id}_pcd'] = 0
            self.image_errors[f'{radar_id}_img'] = 0
            self.image_errors[f'{radar_id}_heatmap'] = 0

    def process(self):
        if not HAS_ROS:
            print("❌ ROS2 라이브러리가 필요합니다.")
            return
            
        print(f"🚀 Processing Bag: {self.bag_path}")
        print(f"📂 Output Directory: {self.output_dir}")

        storage_options = StorageOptions(uri=self.bag_path, storage_id='sqlite3')
        converter_options = ConverterOptions(
            input_serialization_format='cdr', 
            output_serialization_format='cdr'
        )
        reader = SequentialReader()
        reader.open(storage_options, converter_options)

        topics = reader.get_all_topics_and_types()
        type_map = {topic.name: topic.type for topic in topics}
        
        print("\n📋 발견된 토픽:")
        for topic in topics:
            print(f"   - {topic.name}: {topic.type}")
        print()

        while reader.has_next():
            (topic, data, t) = reader.read_next()
            
            try:
                # 멀티 레이더 개별 (radar_0, radar_1, ...)
                radar_match = re.match(r'/sync/(radar_\d+)/(\w+)', topic)
                if radar_match:
                    radar_id = radar_match.group(1)
                    data_type = radar_match.group(2)
                    self.handle_radar_individual(radar_id, data_type, data, t)
                    continue
                
                # merged 레이더
                if '/sync/radar/merged/points' in topic:
                    self.handle_radar_merged(data, t)
                elif '/sync/camera/color' in topic:
                    self.handle_camera(data, t)
                elif '/sync/audio/spectrum' in topic:
                    self.handle_audio_spectrum(data, t)
                elif '/sync/audio/direction' in topic:
                    self.handle_audio_direction(data, t)
                elif '/sync/frame_info' in topic:
                    self.handle_frame_info(data)
                    
            except Exception as e:
                print(f"   ⚠️ 처리 오류 ({topic}): {e}")
                traceback.print_exc()
                continue

        self.save_csvs()
        self.print_summary()

    # ==========================================
    # 멀티 레이더 개별 처리
    # ==========================================
    def handle_radar_individual(self, radar_id, data_type, data, t):
        self.setup_radar_dirs(radar_id)
        timestamp = t / 1e9
        
        if data_type == 'points':
            msg = deserialize_message(data, PointCloud2)
            msg_timestamp = msg.header.stamp.sec + msg.header.stamp.nanosec / 1e9
            
            self.radar_seq_counts[radar_id] += 1
            frame_idx = self.radar_seq_counts[radar_id]
            
            points = []
            for p in point_cloud2.read_points(msg, field_names=("x", "y", "z", "velocity", "intensity"), skip_nans=True):
                x, y, z, v, intensity = p
                range_m = np.sqrt(x**2 + y**2 + z**2)
                if intensity < 100 and range_m > 0.1:
                    rcs = intensity + 40 * np.log10(range_m)
                else:
                    rcs = intensity
                points.append([x, y, z, v, rcs])
            
            if points:
                self.save_radar_frame_csv(radar_id, points, frame_idx, msg_timestamp)
                self.save_radar_pcd_plot(radar_id, points, frame_idx, msg_timestamp)
                self.save_radar_img_plot(radar_id, points, frame_idx, msg_timestamp)
        
        elif data_type == 'heatmap':
            msg = deserialize_message(data, Image)
            msg_timestamp = msg.header.stamp.sec + msg.header.stamp.nanosec / 1e9
            # 히트맵 전용 카운터 사용 (points 카운터와 독립)
            heatmap_key = f'{radar_id}_heatmap_seq'
            if heatmap_key not in self.radar_seq_counts:
                self.radar_seq_counts[heatmap_key] = 0
            self.radar_seq_counts[heatmap_key] += 1
            frame_idx = self.radar_seq_counts[heatmap_key]
            
            try:
                img = np.frombuffer(msg.data, dtype=np.uint8).reshape(msg.height, msg.width)
                self.save_radar_heatmap_plot(radar_id, img, frame_idx, msg_timestamp)
            except Exception as e:
                self.image_errors[f'{radar_id}_heatmap'] += 1
        
        elif data_type == 'range_profile':
            msg = deserialize_message(data, Float32MultiArray)
            self.radar_data[radar_id]['range'].append([timestamp] + list(msg.data))
        
        elif data_type == 'noise_profile':
            msg = deserialize_message(data, Float32MultiArray)
            self.radar_data[radar_id]['noise'].append([timestamp] + list(msg.data))

    def handle_radar_merged(self, data, t):
        """merged 포인트클라우드 처리"""
        self.setup_radar_dirs('merged')
        
        msg = deserialize_message(data, PointCloud2)
        msg_timestamp = msg.header.stamp.sec + msg.header.stamp.nanosec / 1e9
        
        self.radar_seq_counts['merged'] += 1
        frame_idx = self.radar_seq_counts['merged']
        
        points = []
        for p in point_cloud2.read_points(msg, field_names=("x", "y", "z", "velocity", "intensity"), skip_nans=True):
            x, y, z, v, intensity = p
            range_m = np.sqrt(x**2 + y**2 + z**2)
            if intensity < 100 and range_m > 0.1:
                rcs = intensity + 40 * np.log10(range_m)
            else:
                rcs = intensity
            points.append([x, y, z, v, rcs])
        
        if points:
            self.save_radar_frame_csv('merged', points, frame_idx, msg_timestamp)
            self.save_radar_pcd_plot('merged', points, frame_idx, msg_timestamp)
            self.save_radar_img_plot('merged', points, frame_idx, msg_timestamp)

    # ==========================================
    # 레이더 저장 함수들
    # ==========================================
    def save_radar_frame_csv(self, radar_id, points, frame_idx, timestamp):
        csv_path = os.path.join(self.radar_dirs[radar_id]['csv'], f'frame_{frame_idx:04d}.csv')
        df = pd.DataFrame(points, columns=['x', 'y', 'z', 'velocity', 'rcs'])
        df.insert(0, 'timestamp', timestamp)
        df.insert(0, 'frame', frame_idx)
        df.to_csv(csv_path, index=False)

    def save_radar_pcd_plot(self, radar_id, points, frame_idx, timestamp):
        try:
            pts = np.array(points)
            if len(pts) == 0:
                return
            
            fig, ax = plt.subplots(figsize=(8, 8))
            
            plot_x = -pts[:, 1]
            plot_y = pts[:, 0]
            velocities = pts[:, 3]
            
            fast_mask = np.abs(velocities) >= VELOCITY_THRESHOLD
            slow_mask = ~fast_mask
            
            # 정적 포인트 (파란색 점만)
            if np.any(slow_mask):
                ax.scatter(plot_x[slow_mask], plot_y[slow_mask], 
                          c='blue', s=15, alpha=0.7, label='static_points')
            
            # 동적 포인트 (빨간색 점만, 화살표 없음)
            if np.any(fast_mask):
                ax.scatter(plot_x[fast_mask], plot_y[fast_mask], 
                          c='red', s=15, alpha=0.7, label='Moving objects')
            
            ax.set_xlim(-10, 20)
            ax.set_ylim(0, 30)
            ax.set_xlabel('X (m)')
            ax.set_ylabel('Y (m)')
            ax.set_title(f'{radar_id} PCD - Frame {frame_idx}\nTime: {timestamp:.3f}s')
            ax.grid(True, alpha=0.3)
            ax.legend(loc='lower right', fontsize=8)
            
            plt.savefig(os.path.join(self.radar_dirs[radar_id]['pcd'], f'frame_{frame_idx:04d}.png'), 
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
            
        except Exception as e:
            self.image_errors[f'{radar_id}_pcd'] = self.image_errors.get(f'{radar_id}_pcd', 0) + 1

    def save_radar_img_plot(self, radar_id, points, frame_idx, timestamp):
        try:
            pts = np.array(points)
            if len(pts) == 0:
                return
            
            fig, ax = plt.subplots(figsize=(8, 8))
            
            plot_x = -pts[:, 1]
            plot_y = pts[:, 0]
            velocities = pts[:, 3]
            intensities = pts[:, 4]
            
            fast_mask = np.abs(velocities) >= VELOCITY_THRESHOLD
            slow_mask = ~fast_mask
            
            # 정적 포인트: intensity 기준 컬러맵 (청록~파랑 계열)
            if np.any(slow_mask):
                ax.scatter(plot_x[slow_mask], plot_y[slow_mask], 
                          c=intensities[slow_mask], cmap='winter', s=15, alpha=0.7,
                          label='objects')
            
            # 동적 포인트: 빨간색 점 + 얇은 화살표
            if np.any(fast_mask):
                ax.scatter(plot_x[fast_mask], plot_y[fast_mask],
                          c='red', s=20, alpha=0.8, marker='s', label='moving objects')
                # 속도 벡터 (얇은 화살표, 작은 헤드)
                norms = np.sqrt(plot_x[fast_mask]**2 + plot_y[fast_mask]**2)
                norms[norms == 0] = 1
                u = (plot_x[fast_mask] / norms) * velocities[fast_mask] * 0.3
                v = (plot_y[fast_mask] / norms) * velocities[fast_mask] * 0.3
                ax.quiver(plot_x[fast_mask], plot_y[fast_mask], u, v, 
                         color='red', scale=1, scale_units='xy', angles='xy', 
                         width=0.003, headwidth=2, headlength=2, headaxislength=1.5)
            
            # 최대 속도 계산
            max_vel = np.max(np.abs(velocities)) if len(velocities) > 0 else 0
            
            ax.set_xlim(-10, 20)
            ax.set_ylim(0, 30)
            ax.set_xlabel('X')
            ax.set_ylabel('Y')
            ax.set_title(f'Point Cloud Data({frame_idx}th_frame, {max_vel}m/s)')
            ax.grid(True, alpha=0.3)
            ax.legend(loc='upper right', fontsize=8, markerscale=1.5)
            
            plt.savefig(os.path.join(self.radar_dirs[radar_id]['img'], f'frame_{frame_idx:04d}.png'), 
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
            
        except Exception as e:
            self.image_errors[f'{radar_id}_img'] = self.image_errors.get(f'{radar_id}_img', 0) + 1

    def save_radar_heatmap_plot(self, radar_id, img, frame_idx, timestamp):
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            im = ax.imshow(img, aspect='auto', origin='lower', cmap='jet',
                          interpolation='bicubic', 
                          extent=[0, VIRTUAL_ANTENNAS, 0, MAX_RANGE_METERS])
            
            ax.set_title(f'{radar_id} Heatmap - Frame {frame_idx}\nTime: {timestamp:.3f}s')
            ax.set_xlabel('Azimuth (Virtual Antennas)')
            ax.set_ylabel('Distance (m)')
            plt.colorbar(im, ax=ax, label='Intensity')
            
            plt.savefig(os.path.join(self.radar_dirs[radar_id]['heatmap'], f'frame_{frame_idx:04d}.png'),
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
            
        except Exception as e:
            self.image_errors[f'{radar_id}_heatmap'] = self.image_errors.get(f'{radar_id}_heatmap', 0) + 1

    # ==========================================
    # 카메라 처리
    # ==========================================
    def handle_camera(self, data, t):
        try:
            msg = deserialize_message(data, Image)
            msg_timestamp = msg.header.stamp.sec + msg.header.stamp.nanosec / 1e9
            
            self.seq_counts['camera'] += 1
            frame_idx = self.seq_counts['camera']
            
            encoding = msg.encoding
            
            if encoding in ['rgb8', 'bgr8']:
                n_channels = 3
                dtype = np.uint8
            elif encoding in ['rgba8', 'bgra8']:
                n_channels = 4
                dtype = np.uint8
            elif encoding == 'mono8':
                n_channels = 1
                dtype = np.uint8
            else:
                n_channels = 3
                dtype = np.uint8
            
            if n_channels == 1:
                img = np.frombuffer(msg.data, dtype=dtype).reshape(msg.height, msg.width)
            else:
                img = np.frombuffer(msg.data, dtype=dtype).reshape(msg.height, msg.width, n_channels)
            
            if encoding == 'rgb8':
                img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            elif encoding == 'rgba8':
                img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
            
            cv2.imwrite(os.path.join(self.dirs['camera_images'], f'frame_{frame_idx:04d}.jpg'), img)
                
        except Exception as e:
            self.image_errors['camera'] = self.image_errors.get('camera', 0) + 1
            print(f"   ⚠️ 카메라 오류: {e}")

    # ==========================================
    # 오디오 처리
    # ==========================================
    def handle_audio_spectrum(self, data, t):
        msg = deserialize_message(data, Float32MultiArray)
        timestamp = t / 1e9
        
        self.seq_counts['audio'] += 1
        frame_idx = self.seq_counts['audio']
        spectrum = np.array(msg.data)
        
        num_angles = len(spectrum)
        azimuth_grid = np.linspace(-180, 180, num_angles, endpoint=False)
        
        max_idx = np.argmax(spectrum)
        max_angle = azimuth_grid[max_idx]
        
        self.save_audio_frame_csv(spectrum, azimuth_grid, frame_idx, timestamp, max_angle)
        self.save_audio_plot(spectrum, azimuth_grid, frame_idx, max_angle, timestamp)

    def handle_audio_direction(self, data, t):
        pass  # direction은 spectrum과 함께 처리

    def save_audio_frame_csv(self, spectrum, azimuth_grid, frame_idx, timestamp, max_angle):
        csv_path = os.path.join(self.dirs['acoustic_csv'], f'frame_{frame_idx:04d}.csv')
        df = pd.DataFrame({
            'angle': azimuth_grid,
            'spectrum': spectrum
        })
        df.insert(0, 'max_angle', max_angle)
        df.insert(0, 'timestamp', timestamp)
        df.insert(0, 'frame', frame_idx)
        df.to_csv(csv_path, index=False)

    def save_audio_plot(self, spectrum, azimuth_grid, frame_idx, max_angle, timestamp):
        try:
            fig = plt.figure(figsize=(8, 6))
            ax = fig.add_subplot(111, projection='polar')
            
            ax.set_theta_zero_location('N')
            ax.set_theta_direction(-1)
            
            mask = (azimuth_grid >= AUDIO_ANGLE_MIN) & (azimuth_grid <= AUDIO_ANGLE_MAX)
            angles_filtered = azimuth_grid[mask]
            spectrum_filtered = spectrum[mask]
            
            ax.plot(np.deg2rad(angles_filtered), spectrum_filtered, color='red', linewidth=1.5)
            ax.fill(np.deg2rad(angles_filtered), spectrum_filtered, color='red', alpha=0.3)
            
            if AUDIO_ANGLE_MIN <= max_angle <= AUDIO_ANGLE_MAX:
                max_val = spectrum[np.argmax(spectrum)]
                ax.plot(np.deg2rad(max_angle), max_val, 'bo', markersize=10)
            
            ax.set_thetamin(AUDIO_ANGLE_MIN)
            ax.set_thetamax(AUDIO_ANGLE_MAX)
            ax.set_title(f'Audio - Frame {frame_idx} | Max: {max_angle:.1f}°\nTime: {timestamp:.3f}s')
            
            plt.savefig(os.path.join(self.dirs['acoustic_plots'], f'frame_{frame_idx:04d}.png'),
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
            
        except Exception as e:
            self.image_errors['audio'] = self.image_errors.get('audio', 0) + 1

    def handle_frame_info(self, data):
        msg = deserialize_message(data, String)
        self.data_store['frame_info'].append(msg.data)

    # ==========================================
    # CSV 저장
    # ==========================================
    def save_csvs(self):
        print("\n💾 Saving CSV files...")
        
        # 레이더별 range/noise profile (프레임 번호 추가)
        for radar_id, rdata in self.radar_data.items():
            if radar_id == 'merged':
                continue
                
            if rdata['range']:
                num_bins = len(rdata['range'][0]) - 1
                columns = ['timestamp'] + [f'bin_{i}' for i in range(num_bins)]
                df = pd.DataFrame(rdata['range'], columns=columns)
                df.insert(0, 'frame', range(1, len(df) + 1))
                save_path = os.path.join(self.dirs['radar'], radar_id, 'range_profile.csv')
                df.to_csv(save_path, index=False)
                print(f"   ✅ {radar_id}/range_profile.csv ({len(df)} frames)")
            
            if rdata['noise']:
                num_bins = len(rdata['noise'][0]) - 1
                columns = ['timestamp'] + [f'bin_{i}' for i in range(num_bins)]
                df = pd.DataFrame(rdata['noise'], columns=columns)
                df.insert(0, 'frame', range(1, len(df) + 1))
                save_path = os.path.join(self.dirs['radar'], radar_id, 'noise_profile.csv')
                df.to_csv(save_path, index=False)
                print(f"   ✅ {radar_id}/noise_profile.csv ({len(df)} frames)")

        # sync_quality.xlsx (엑셀 - 노란색 하이라이트 포함)
        if self.data_store['frame_info']:
            self.save_sync_quality_xlsx()

    def save_sync_quality_xlsx(self):
        """sync_quality를 xlsx로 저장 (hold 행 노란색 하이라이트)"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        rows = []
        for info_str in self.data_store['frame_info']:
            parts = info_str.split(',')
            if len(parts) >= 5:
                rows.append(parts)
        
        if not rows:
            return
        
        # 컬럼 수에서 레이더 수 추정
        # format: frame,sync_time,sensors_new,sensors_held,max_diff,
        #         [radar_i_orig,radar_i_diff]*N, cam_orig,cam_diff,cam_held,
        #         aud_orig,aud_diff,aud_held, lidar_orig,lidar_diff
        sample_len = len(rows[0])
        # 고정: 5(기본) + 3(cam) + 3(aud) + 2(lidar) = 13, 나머지는 레이더*2
        num_radar_cols = sample_len - 13
        num_radars = num_radar_cols // 2 if num_radar_cols > 0 else 0
        
        # 헤더 생성
        headers = ['frame', 'sync_time', 'sensors_new', 'sensors_held', 'max_diff_ms']
        for i in range(num_radars):
            headers += [f'radar_{i}_orig', f'radar_{i}_diff_ms']
        headers += ['camera_orig', 'camera_diff_ms', 'camera_held',
                    'audio_orig', 'audio_diff_ms', 'audio_held',
                    'lidar_orig', 'lidar_diff_ms']
        
        # 실제 열 수에 맞춰 조정
        headers = headers[:sample_len]
        
        # hold 열 인덱스 찾기 (camera_held, audio_held)
        cam_held_idx = None
        aud_held_idx = None
        for i, h in enumerate(headers):
            if h == 'camera_held':
                cam_held_idx = i
            elif h == 'audio_held':
                aud_held_idx = i
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'sync_quality'
        
        # 스타일
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='D9E1F2')
        yellow_fill = PatternFill('solid', fgColor='FFFF00')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center')
        
        # 헤더 쓰기
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 데이터 쓰기
        hold_rows = []
        for row_idx, parts in enumerate(rows, 2):
            seq_frame = row_idx - 1
            is_hold = False
            
            for col_idx, val in enumerate(parts):
                # 숫자 변환 시도
                try:
                    if '.' in val:
                        cell_val = float(val)
                    else:
                        cell_val = int(val)
                except ValueError:
                    cell_val = val
                
                # 첫 열은 순차 프레임 번호로 교체
                if col_idx == 0:
                    cell_val = seq_frame
                
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=cell_val)
                cell.border = thin_border
                cell.alignment = center_align
                
                # hold 여부 체크
                if col_idx == cam_held_idx or col_idx == aud_held_idx:
                    try:
                        if int(float(val)) == 1:
                            is_hold = True
                    except:
                        pass
            
            if is_hold:
                hold_rows.append(row_idx)
        
        # hold 행 노란색 칠하기
        for row_idx in hold_rows:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
        
        # 열 너비 자동 조정
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 2, 12)
        
        # 필터 추가
        ws.auto_filter.ref = ws.dimensions
        
        # 첫 행 고정
        ws.freeze_panes = 'A2'
        
        save_path = os.path.join(self.output_dir, 'sync_quality.xlsx')
        wb.save(save_path)
        print(f"   ✅ sync_quality.xlsx ({len(rows)} frames, {len(hold_rows)} hold rows highlighted)")

    def print_summary(self):
        print("\n" + "="*60)
        print("                    처리 완료 요약")
        print("="*60)
        print(f"  📁 출력 경로: {self.output_dir}")
        
        # 레이더별 프레임 수
        for radar_id, count in sorted(self.radar_seq_counts.items()):
            print(f"  📊 {radar_id}: {count} frames")
        
        print(f"  📊 카메라: {self.seq_counts['camera']} frames")
        print(f"  📊 오디오: {self.seq_counts['audio']} frames")
        print()
        
        # 이미지 생성 오류
        total_errors = sum(self.image_errors.values())
        if total_errors > 0:
            print("  ⚠️ 이미지 생성 오류:")
            for key, count in self.image_errors.items():
                if count > 0:
                    print(f"     - {key}: {count}개")
        else:
            print("  ✅ 모든 이미지 정상 생성")
        
        print("="*60)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='NLOS Multi-Sensor Rosbag Processor v2')
    parser.add_argument('bag_path', nargs='?', default=None)
    
    args = parser.parse_args()
    
    if args.bag_path:
        bag_path = os.path.expanduser(args.bag_path)
    else:
        bag_path = get_latest_bag_path(ROSBAG_DIR)
    
    if bag_path and os.path.exists(bag_path):
        processor = BagProcessor(bag_path)
        processor.process()
    else:
        print("❌ rosbag을 찾을 수 없습니다.")
