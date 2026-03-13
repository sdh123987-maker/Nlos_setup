#!/usr/bin/env python3
"""
==============================================
NLOS Multi-Sensor Rosbag Processor v3
==============================================
v6 (Raw Recording) 구조에 맞춤

핵심 변경점 (v2 → v3):
  - 실시간 동기화 노드 없이, raw rosbag에서 직접 10Hz 동기화 수행
  - 1단계: 모든 메시지를 토픽별로 타임스탬프와 함께 수집
  - 2단계: 10Hz 그리드 생성 → 각 센서별 nearest-neighbor 매칭
  - 라이다: Ouster 네이티브 PCD 포맷(binary) 저장 추가
  - 오디오: raw 음원 데이터를 WAV 파일로 저장 추가

토픽 구조 (raw - /sync/ 접두사 없음):
  /radar/merged/points         - 통합 포인트클라우드
  /radar_0/points              - 레이더0 개별
  /radar_0/heatmap/azimuth     - 레이더0 히트맵
  /radar_1/points              - 레이더1 개별
  /camera/camera/color/image_raw
  /audio/spectrum, /audio/direction, /audio/raw
  /ouster/points, /ouster/imu

사용법:
  python3 process_rosbag.py                    # 최신 rosbag 자동 처리
  python3 process_rosbag.py /path/to/rosbag    # 특정 rosbag 처리
  python3 process_rosbag.py --sync-rate 20     # 20Hz 동기화
"""

import os
import sys
import glob
import struct
import cv2
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
import traceback
import warnings
import re
import wave
warnings.filterwarnings('ignore')

try:
    from rclpy.serialization import deserialize_message
    from rosbag2_py import SequentialReader, StorageOptions, ConverterOptions
    from sensor_msgs.msg import PointCloud2, Image, Imu
    from std_msgs.msg import Float32MultiArray, Float32, String, Header
    from sensor_msgs_py import point_cloud2
    HAS_ROS = True
except ImportError:
    HAS_ROS = False
    print("⚠️ ROS2 라이브러리가 필요합니다.")

# ==========================================
# 설정
# ==========================================
WORKSPACE_DIR = os.path.expanduser('~/Ros2_setup')
ROSBAG_DIR = os.path.join(WORKSPACE_DIR, 'rosbag')
OUTPUT_BASE_DIR = os.path.join(WORKSPACE_DIR, 'processed_data')

MAX_RANGE_METERS = 10.0
VIRTUAL_ANTENNAS = 12
VELOCITY_THRESHOLD = 0.1

AUDIO_ANGLE_MIN, AUDIO_ANGLE_MAX = 0, 180

# 오디오 설정
AUDIO_CHANNELS = 16
AUDIO_SAMPLE_RATE = 48000


def get_latest_bag_path(base_dir):
    all_bags = glob.glob(os.path.join(base_dir, 'raw_*')) + glob.glob(os.path.join(base_dir, 'sync_*'))
    if not all_bags:
        print("❌ Rosbag 폴더가 비어있습니다.")
        return None
    return max(all_bags, key=os.path.getctime)


def msg_timestamp_sec(t_ns):
    """rosbag 타임스탬프(nanosecond) → seconds"""
    return t_ns / 1e9


def header_timestamp_sec(msg):
    """ROS Header → seconds"""
    return msg.header.stamp.sec + msg.header.stamp.nanosec / 1e9


class TimestampedMessage:
    """타임스탬프 + 역직렬화된 메시지를 함께 저장"""
    __slots__ = ['timestamp', 'msg']
    def __init__(self, timestamp, msg):
        self.timestamp = timestamp
        self.msg = msg


class BagProcessor:
    def __init__(self, bag_path, sync_rate_hz=10.0):
        self.bag_path = bag_path
        self.bag_name = os.path.basename(bag_path)
        self.output_dir = os.path.join(OUTPUT_BASE_DIR, self.bag_name)
        self.sync_rate_hz = sync_rate_hz
        
        # 기본 폴더
        self.dirs = {
            'radar': os.path.join(self.output_dir, 'radar'),
            'lidar': os.path.join(self.output_dir, 'lidar'),
            'lidar_csv': os.path.join(self.output_dir, 'lidar', 'csv'),
            'lidar_pcd': os.path.join(self.output_dir, 'lidar', 'pcd'),
            'lidar_img': os.path.join(self.output_dir, 'lidar', 'img'),
            'camera': os.path.join(self.output_dir, 'camera'),
            'camera_images': os.path.join(self.output_dir, 'camera', 'images'),
            'acoustic': os.path.join(self.output_dir, 'acoustic'),
            'acoustic_csv': os.path.join(self.output_dir, 'acoustic', 'csv'),
            'acoustic_plots': os.path.join(self.output_dir, 'acoustic', 'plots'),
            'acoustic_wav': os.path.join(self.output_dir, 'acoustic', 'wav'),
        }
        
        self.radar_dirs = {}
        
        for d in self.dirs.values():
            os.makedirs(d, exist_ok=True)

        # ============================================================
        # 1단계 데이터 수집: 토픽별로 (timestamp, deserialized_msg) 리스트
        # ============================================================
        self.raw_messages = {}  # topic_name -> [TimestampedMessage, ...]
        
        # 오디오 raw 데이터 버퍼 (WAV 저장용)
        self.audio_raw_chunks = []  # [(timestamp, np.array), ...]
        
        # 레이더별 폴더
        self.radar_ids = set()
        
        # 이미지 생성 오류 추적
        self.image_errors = {}
        
        # 통계
        self.sync_stats = {}

    def setup_radar_dirs(self, radar_id):
        if radar_id not in self.radar_dirs:
            base = os.path.join(self.dirs['radar'], radar_id)
            self.radar_dirs[radar_id] = {
                'csv': os.path.join(base, 'csv'),
                'pcd': os.path.join(base, 'pcd'),
                'img': os.path.join(base, 'img'),
                'heatmap': os.path.join(base, 'heatmap'),
            }
            for d in self.radar_dirs[radar_id].values():
                os.makedirs(d, exist_ok=True)
            self.radar_ids.add(radar_id)

    def process(self):
        if not HAS_ROS:
            print("❌ ROS2 라이브러리가 필요합니다.")
            return
            
        print(f"🚀 Processing Bag: {self.bag_path}")
        print(f"📂 Output Directory: {self.output_dir}")
        print(f"🔄 Sync Rate: {self.sync_rate_hz} Hz")
        print()

        # ============================================================
        # 1단계: 모든 메시지 수집
        # ============================================================
        print("=" * 60)
        print("  1단계: Raw 메시지 수집")
        print("=" * 60)
        self.collect_all_messages()
        
        # ============================================================
        # 2단계: 10Hz 동기화 + 데이터 추출/저장
        # ============================================================
        print()
        print("=" * 60)
        print("  2단계: 후처리 동기화 + 데이터 저장")
        print("=" * 60)
        self.synchronize_and_save()
        
        # ============================================================
        # 3단계: 오디오 raw WAV 저장 (동기화와 무관하게 전체 저장)
        # ============================================================
        self.save_audio_wav()
        
        self.print_summary()

    # ==========================================
    # 1단계: 메시지 수집
    # ==========================================
    def collect_all_messages(self):
        storage_options = StorageOptions(uri=self.bag_path, storage_id='sqlite3')
        converter_options = ConverterOptions(
            input_serialization_format='cdr', 
            output_serialization_format='cdr'
        )
        reader = SequentialReader()
        reader.open(storage_options, converter_options)

        topics = reader.get_all_topics_and_types()
        type_map = {topic.name: topic.type for topic in topics}
        
        print("\n📋 발견된 토픽:")
        for topic in topics:
            print(f"   - {topic.name}: {topic.type}")
        print()

        # 관심 토픽 패턴 (raw 토픽)
        msg_count = 0
        while reader.has_next():
            (topic, data, t) = reader.read_next()
            timestamp = msg_timestamp_sec(t)
            
            try:
                msg_type_str = type_map.get(topic, '')
                
                # 역직렬화할 토픽만 처리
                msg = None
                
                # 레이더 개별
                radar_match = re.match(r'/radar_(\d+)/(points|heatmap/azimuth|range_profile|noise_profile|timestamp)', topic)
                if radar_match:
                    radar_id = f'radar_{radar_match.group(1)}'
                    self.setup_radar_dirs(radar_id)
                    data_type = radar_match.group(2)
                    
                    if data_type == 'points':
                        msg = deserialize_message(data, PointCloud2)
                        timestamp = header_timestamp_sec(msg)
                    elif data_type == 'heatmap/azimuth':
                        msg = deserialize_message(data, Image)
                        timestamp = header_timestamp_sec(msg)
                    elif data_type == 'range_profile':
                        msg = deserialize_message(data, Float32MultiArray)
                    elif data_type == 'noise_profile':
                        msg = deserialize_message(data, Float32MultiArray)
                
                # 레이더 merged
                elif topic == '/radar/merged/points':
                    self.setup_radar_dirs('merged')
                    msg = deserialize_message(data, PointCloud2)
                    timestamp = header_timestamp_sec(msg)
                
                # 카메라
                elif topic == '/camera/camera/color/image_raw':
                    msg = deserialize_message(data, Image)
                    timestamp = header_timestamp_sec(msg)
                
                # 오디오
                elif topic == '/audio/spectrum':
                    msg = deserialize_message(data, Float32MultiArray)
                elif topic == '/audio/direction':
                    msg = deserialize_message(data, Float32)
                elif topic == '/audio/raw':
                    msg = deserialize_message(data, Float32MultiArray)
                    # raw 오디오는 별도 버퍼에도 저장 (WAV용)
                    raw_data = np.array(msg.data, dtype=np.float32)
                    self.audio_raw_chunks.append((timestamp, raw_data))
                elif topic == '/audio/header':
                    msg = deserialize_message(data, Header)
                    timestamp = msg.stamp.sec + msg.stamp.nanosec / 1e9
                
                # 라이다
                elif topic == '/ouster/points':
                    msg = deserialize_message(data, PointCloud2)
                    timestamp = header_timestamp_sec(msg)
                elif topic == '/ouster/imu':
                    msg = deserialize_message(data, Imu)
                    timestamp = header_timestamp_sec(msg)
                elif topic == '/ouster/range_image':
                    msg = deserialize_message(data, Image)
                    timestamp = header_timestamp_sec(msg)
                
                if msg is not None:
                    if topic not in self.raw_messages:
                        self.raw_messages[topic] = []
                    self.raw_messages[topic].append(TimestampedMessage(timestamp, msg))
                    msg_count += 1
                    
            except Exception as e:
                pass  # 역직렬화 실패 무시
        
        # 수집 결과
        print(f"\n📊 수집 완료: 총 {msg_count} 메시지")
        for topic, msgs in sorted(self.raw_messages.items()):
            t_min = msgs[0].timestamp if msgs else 0
            t_max = msgs[-1].timestamp if msgs else 0
            print(f"   {topic}: {len(msgs)} msgs ({t_max - t_min:.1f}s)")

    # ==========================================
    # 2단계: 10Hz 동기화 + 저장
    # ==========================================
    def synchronize_and_save(self):
        # 전체 시간 범위 계산
        all_timestamps = []
        for msgs in self.raw_messages.values():
            for m in msgs:
                all_timestamps.append(m.timestamp)
        
        if not all_timestamps:
            print("❌ 처리할 데이터가 없습니다.")
            return
        
        t_start = min(all_timestamps)
        t_end = max(all_timestamps)
        duration = t_end - t_start
        
        # 10Hz 동기화 그리드 생성
        period = 1.0 / self.sync_rate_hz
        sync_times = np.arange(t_start + period, t_end, period)  # 첫 프레임은 약간 여유
        num_frames = len(sync_times)
        
        print(f"\n⏱️  시간 범위: {t_start:.3f} ~ {t_end:.3f} ({duration:.1f}s)")
        print(f"📐 동기화 프레임: {num_frames} frames @ {self.sync_rate_hz}Hz")
        print()
        
        # 각 토픽에 대해 numpy 타임스탬프 배열 준비 (bisect용)
        topic_timestamps = {}
        for topic, msgs in self.raw_messages.items():
            topic_timestamps[topic] = np.array([m.timestamp for m in msgs])
        
        # sync_quality 데이터
        sync_quality_rows = []
        
        # 센서별 카운터
        frame_counts = {'radar_merged': 0, 'camera': 0, 'audio': 0, 'lidar': 0}
        for rid in self.radar_ids:
            if rid != 'merged':
                frame_counts[rid] = 0
        
        # ============================================================
        # 메인 동기화 루프
        # ============================================================
        for frame_idx, sync_t in enumerate(sync_times, start=1):
            quality_row = {
                'frame': frame_idx,
                'sync_time': sync_t,
            }
            
            # --- 레이더 merged ---
            self._process_synced_radar(
                'merged', sync_t, frame_idx, topic_timestamps, quality_row, frame_counts)
            
            # --- 레이더 개별 ---
            for radar_id in sorted(self.radar_ids):
                if radar_id == 'merged':
                    continue
                self._process_synced_radar(
                    radar_id, sync_t, frame_idx, topic_timestamps, quality_row, frame_counts)
            
            # --- 카메라 ---
            cam_topic = '/camera/camera/color/image_raw'
            if cam_topic in self.raw_messages:
                nearest_msg, diff_ms = self._find_nearest(cam_topic, sync_t, topic_timestamps)
                quality_row['camera_diff_ms'] = diff_ms
                if nearest_msg is not None:
                    frame_counts['camera'] += 1
                    cam_idx = frame_counts['camera']
                    self._save_camera_image(nearest_msg.msg, cam_idx)
            
            # --- 라이다 ---
            lidar_topic = '/ouster/points'
            if lidar_topic in self.raw_messages:
                nearest_msg, diff_ms = self._find_nearest(lidar_topic, sync_t, topic_timestamps)
                quality_row['lidar_diff_ms'] = diff_ms
                if nearest_msg is not None:
                    frame_counts['lidar'] += 1
                    lid_idx = frame_counts['lidar']
                    self._save_lidar_frame(nearest_msg.msg, lid_idx, nearest_msg.timestamp)
            
            # --- 라이다 IMU ---
            imu_topic = '/ouster/imu'
            if imu_topic in self.raw_messages:
                nearest_msg, _ = self._find_nearest(imu_topic, sync_t, topic_timestamps)
                # IMU는 CSV로 나중에 일괄 저장
            
            # --- 오디오 spectrum ---
            spec_topic = '/audio/spectrum'
            if spec_topic in self.raw_messages:
                nearest_msg, diff_ms = self._find_nearest(spec_topic, sync_t, topic_timestamps)
                quality_row['audio_diff_ms'] = diff_ms
                if nearest_msg is not None:
                    frame_counts['audio'] += 1
                    aud_idx = frame_counts['audio']
                    self._save_audio_spectrum(nearest_msg.msg, aud_idx, nearest_msg.timestamp)
            
            sync_quality_rows.append(quality_row)
            
            # 진행률
            if frame_idx % 100 == 0 or frame_idx == num_frames:
                print(f"   처리중: {frame_idx}/{num_frames} frames ({frame_idx/num_frames*100:.0f}%)")
        
        # ============================================================
        # 일괄 CSV 저장
        # ============================================================
        self._save_bulk_csvs(topic_timestamps)
        self._save_sync_quality(sync_quality_rows)
        
        self.sync_stats = frame_counts

    def _find_nearest(self, topic, target_time, topic_timestamps):
        """주어진 토픽에서 target_time에 가장 가까운 메시지 찾기"""
        if topic not in self.raw_messages or len(self.raw_messages[topic]) == 0:
            return None, float('inf')
        
        ts = topic_timestamps[topic]
        idx = np.searchsorted(ts, target_time)
        
        # 양쪽 후보 비교
        best_idx = None
        best_diff = float('inf')
        
        for candidate in [idx - 1, idx]:
            if 0 <= candidate < len(ts):
                diff = abs(ts[candidate] - target_time)
                if diff < best_diff:
                    best_diff = diff
                    best_idx = candidate
        
        if best_idx is not None:
            diff_ms = best_diff * 1000.0
            return self.raw_messages[topic][best_idx], diff_ms
        
        return None, float('inf')

    # ==========================================
    # 레이더 처리
    # ==========================================
    def _process_synced_radar(self, radar_id, sync_t, frame_idx, topic_timestamps, quality_row, frame_counts):
        if radar_id == 'merged':
            pts_topic = '/radar/merged/points'
        else:
            pts_topic = f'/{radar_id}/points'
        
        if pts_topic not in self.raw_messages:
            return
        
        nearest_msg, diff_ms = self._find_nearest(pts_topic, sync_t, topic_timestamps)
        quality_row[f'{radar_id}_diff_ms'] = diff_ms
        
        if nearest_msg is None:
            return
        
        frame_counts[radar_id] = frame_counts.get(radar_id, 0) + 1
        ridx = frame_counts[radar_id]
        
        # 포인트클라우드 추출
        msg = nearest_msg.msg
        msg_timestamp = header_timestamp_sec(msg)
        
        points = []
        for p in point_cloud2.read_points(msg, field_names=("x", "y", "z", "velocity", "intensity"), skip_nans=True):
            x, y, z, v, intensity = p
            range_m = np.sqrt(x**2 + y**2 + z**2)
            if intensity < 100 and range_m > 0.1:
                rcs = intensity + 40 * np.log10(range_m)
            else:
                rcs = intensity
            points.append([x, y, z, v, rcs])
        
        if points:
            self._save_radar_frame_csv(radar_id, points, ridx, msg_timestamp)
            self._save_radar_pcd_plot(radar_id, points, ridx, msg_timestamp)
            self._save_radar_img_plot(radar_id, points, ridx, msg_timestamp)
        
        # 히트맵
        if radar_id != 'merged':
            hm_topic = f'/{radar_id}/heatmap/azimuth'
            if hm_topic in self.raw_messages:
                hm_msg, _ = self._find_nearest(hm_topic, sync_t, topic_timestamps)
                if hm_msg is not None:
                    try:
                        img = np.frombuffer(hm_msg.msg.data, dtype=np.uint8).reshape(
                            hm_msg.msg.height, hm_msg.msg.width)
                        self._save_radar_heatmap_plot(radar_id, img, ridx, hm_msg.timestamp)
                    except Exception:
                        pass

    def _save_radar_frame_csv(self, radar_id, points, frame_idx, timestamp):
        self.setup_radar_dirs(radar_id)
        csv_path = os.path.join(self.radar_dirs[radar_id]['csv'], f'frame_{frame_idx:04d}.csv')
        df = pd.DataFrame(points, columns=['x', 'y', 'z', 'velocity', 'rcs'])
        df.insert(0, 'timestamp', timestamp)
        df.insert(0, 'frame', frame_idx)
        df.to_csv(csv_path, index=False)

    def _save_radar_pcd_plot(self, radar_id, points, frame_idx, timestamp):
        try:
            pts = np.array(points)
            if len(pts) == 0:
                return
            
            fig, ax = plt.subplots(figsize=(8, 8))
            plot_x = -pts[:, 1]
            plot_y = pts[:, 0]
            velocities = pts[:, 3]
            
            fast_mask = np.abs(velocities) >= VELOCITY_THRESHOLD
            slow_mask = ~fast_mask
            
            if np.any(slow_mask):
                ax.scatter(plot_x[slow_mask], plot_y[slow_mask], 
                          c='blue', s=15, alpha=0.7, label='static_points')
            if np.any(fast_mask):
                ax.scatter(plot_x[fast_mask], plot_y[fast_mask], 
                          c='red', s=15, alpha=0.7, label='Moving objects')
            
            ax.set_xlim(-10, 20)
            ax.set_ylim(0, 30)
            ax.set_xlabel('X (m)')
            ax.set_ylabel('Y (m)')
            ax.set_title(f'{radar_id} PCD - Frame {frame_idx}\nTime: {timestamp:.3f}s')
            ax.grid(True, alpha=0.3)
            ax.legend(loc='lower right', fontsize=8)
            
            plt.savefig(os.path.join(self.radar_dirs[radar_id]['pcd'], f'frame_{frame_idx:04d}.png'), 
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
        except Exception:
            self.image_errors[f'{radar_id}_pcd'] = self.image_errors.get(f'{radar_id}_pcd', 0) + 1

    def _save_radar_img_plot(self, radar_id, points, frame_idx, timestamp):
        try:
            pts = np.array(points)
            if len(pts) == 0:
                return
            
            fig, ax = plt.subplots(figsize=(8, 8))
            plot_x = -pts[:, 1]
            plot_y = pts[:, 0]
            velocities = pts[:, 3]
            intensities = pts[:, 4]
            
            fast_mask = np.abs(velocities) >= VELOCITY_THRESHOLD
            slow_mask = ~fast_mask
            
            if np.any(slow_mask):
                ax.scatter(plot_x[slow_mask], plot_y[slow_mask], 
                          c=intensities[slow_mask], cmap='winter', s=15, alpha=0.7,
                          label='objects')
            if np.any(fast_mask):
                ax.scatter(plot_x[fast_mask], plot_y[fast_mask],
                          c='red', s=20, alpha=0.8, marker='s', label='moving objects')
                norms = np.sqrt(plot_x[fast_mask]**2 + plot_y[fast_mask]**2)
                norms[norms == 0] = 1
                u = (plot_x[fast_mask] / norms) * velocities[fast_mask] * 0.3
                v = (plot_y[fast_mask] / norms) * velocities[fast_mask] * 0.3
                ax.quiver(plot_x[fast_mask], plot_y[fast_mask], u, v, 
                         color='red', scale=1, scale_units='xy', angles='xy', 
                         width=0.003, headwidth=2, headlength=2, headaxislength=1.5)
            
            max_vel = np.max(np.abs(velocities)) if len(velocities) > 0 else 0
            
            ax.set_xlim(-10, 20)
            ax.set_ylim(0, 30)
            ax.set_xlabel('X')
            ax.set_ylabel('Y')
            ax.set_title(f'Point Cloud Data({frame_idx}th_frame, {max_vel}m/s)')
            ax.grid(True, alpha=0.3)
            ax.legend(loc='upper right', fontsize=8, markerscale=1.5)
            
            plt.savefig(os.path.join(self.radar_dirs[radar_id]['img'], f'frame_{frame_idx:04d}.png'), 
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
        except Exception:
            self.image_errors[f'{radar_id}_img'] = self.image_errors.get(f'{radar_id}_img', 0) + 1

    def _save_radar_heatmap_plot(self, radar_id, img, frame_idx, timestamp):
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            im = ax.imshow(img, aspect='auto', origin='lower', cmap='jet',
                          interpolation='bicubic', 
                          extent=[0, VIRTUAL_ANTENNAS, 0, MAX_RANGE_METERS])
            ax.set_title(f'{radar_id} Heatmap - Frame {frame_idx}\nTime: {timestamp:.3f}s')
            ax.set_xlabel('Azimuth (Virtual Antennas)')
            ax.set_ylabel('Distance (m)')
            plt.colorbar(im, ax=ax, label='Intensity')
            plt.savefig(os.path.join(self.radar_dirs[radar_id]['heatmap'], f'frame_{frame_idx:04d}.png'),
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
        except Exception:
            self.image_errors[f'{radar_id}_heatmap'] = self.image_errors.get(f'{radar_id}_heatmap', 0) + 1

    # ==========================================
    # 카메라 처리
    # ==========================================
    def _save_camera_image(self, msg, frame_idx):
        try:
            encoding = msg.encoding
            if encoding in ['rgb8', 'bgr8']:
                n_channels, dtype = 3, np.uint8
            elif encoding in ['rgba8', 'bgra8']:
                n_channels, dtype = 4, np.uint8
            elif encoding == 'mono8':
                n_channels, dtype = 1, np.uint8
            else:
                n_channels, dtype = 3, np.uint8
            
            if n_channels == 1:
                img = np.frombuffer(msg.data, dtype=dtype).reshape(msg.height, msg.width)
            else:
                img = np.frombuffer(msg.data, dtype=dtype).reshape(msg.height, msg.width, n_channels)
            
            if encoding == 'rgb8':
                img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            elif encoding == 'rgba8':
                img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
            
            cv2.imwrite(os.path.join(self.dirs['camera_images'], f'frame_{frame_idx:04d}.jpg'), img)
        except Exception as e:
            self.image_errors['camera'] = self.image_errors.get('camera', 0) + 1

    # ==========================================
    # 라이다 처리 (PCD binary 포맷 + BEV 이미지 + CSV)
    # ==========================================
    def _save_lidar_frame(self, msg, frame_idx, timestamp):
        """라이다 프레임: PCD(binary) + BEV 이미지 + CSV"""
        # PCD 저장 (Ouster 네이티브 필드 그대로)
        self._save_lidar_pcd(msg, frame_idx)
        
        # CSV + BEV 이미지 (x, y, z, intensity 필터링)
        try:
            points = []
            for p in point_cloud2.read_points(msg, skip_nans=True):
                # Ouster 필드: x, y, z, (padding), intensity, t, reflectivity, ring, ambient, (padding), range, (padding)
                # read_points는 유효 필드만 반환
                x = float(p[0]) if len(p) > 0 else 0.0
                y = float(p[1]) if len(p) > 1 else 0.0
                z = float(p[2]) if len(p) > 2 else 0.0
                intensity = float(p[3]) if len(p) > 3 else 0.0
                
                points.append([x, y, z, intensity])
            
            if points:
                self._save_lidar_csv(points, frame_idx, timestamp)
                self._save_lidar_bev_plot(points, frame_idx, timestamp)
        except Exception as e:
            self.image_errors['lidar_csv'] = self.image_errors.get('lidar_csv', 0) + 1

    def _save_lidar_pcd(self, msg, frame_idx):
        """PointCloud2 → PCD binary 파일 (Ouster 네이티브 포맷 보존)
        
        업로드된 PCD 포맷과 동일:
          FIELDS x y z _ intensity t reflectivity ring ambient _ range _
          DATA binary
        """
        pcd_path = os.path.join(self.dirs['lidar_pcd'], f'frame_{frame_idx:05d}.pcd')
        
        try:
            # PointCloud2에서 필드 정보 추출
            fields_str = ' '.join([f.name for f in msg.fields])
            sizes_str = ' '.join([str(self._datatype_to_size(f.datatype)) for f in msg.fields])
            types_str = ' '.join([self._datatype_to_type_char(f.datatype) for f in msg.fields])
            counts_str = ' '.join([str(f.count) for f in msg.fields])
            
            width = msg.width
            height = msg.height
            num_points = width * height
            
            # PCD 헤더 작성
            header = (
                f"# .PCD v0.7 - Point Cloud Data file format\n"
                f"VERSION 0.7\n"
                f"FIELDS {fields_str}\n"
                f"SIZE {sizes_str}\n"
                f"TYPE {types_str}\n"
                f"COUNT {counts_str}\n"
                f"WIDTH {width}\n"
                f"HEIGHT {height}\n"
                f"VIEWPOINT 0 0 0 1 0 0 0\n"
                f"POINTS {num_points}\n"
                f"DATA binary\n"
            )
            
            with open(pcd_path, 'wb') as f:
                f.write(header.encode('ascii'))
                # raw binary 데이터 그대로 기록
                f.write(bytes(msg.data))
                
        except Exception as e:
            self.image_errors['lidar_pcd'] = self.image_errors.get('lidar_pcd', 0) + 1

    def _datatype_to_size(self, datatype):
        """sensor_msgs/PointField datatype → byte size"""
        size_map = {1: 1, 2: 1, 3: 2, 4: 2, 5: 4, 6: 4, 7: 4, 8: 8}
        return size_map.get(datatype, 4)

    def _datatype_to_type_char(self, datatype):
        """sensor_msgs/PointField datatype → PCD type character"""
        type_map = {1: 'I', 2: 'U', 3: 'I', 4: 'U', 5: 'I', 6: 'U', 7: 'F', 8: 'F'}
        return type_map.get(datatype, 'F')

    def _save_lidar_csv(self, points, frame_idx, timestamp):
        csv_path = os.path.join(self.dirs['lidar_csv'], f'frame_{frame_idx:05d}.csv')
        df = pd.DataFrame(points, columns=['x', 'y', 'z', 'intensity'])
        df.insert(0, 'timestamp', timestamp)
        df.insert(0, 'frame', frame_idx)
        df.to_csv(csv_path, index=False)

    def _save_lidar_bev_plot(self, points, frame_idx, timestamp):
        """라이다 BEV 플롯"""
        try:
            pts = np.array(points)
            if len(pts) == 0:
                return
            
            # 범위 필터링
            x_mask = (pts[:, 0] >= -10) & (pts[:, 0] <= 10)
            y_mask = (pts[:, 1] >= -5) & (pts[:, 1] <= 10)
            mask = x_mask & y_mask
            pts = pts[mask]
            
            if len(pts) == 0:
                return
            
            fig, ax = plt.subplots(figsize=(10, 8))
            scatter = ax.scatter(pts[:, 0], pts[:, 1], c=pts[:, 3], 
                               cmap='viridis', s=0.5, alpha=0.8)
            ax.set_xlim(-10, 10)
            ax.set_ylim(-5, 10)
            ax.set_xlabel('X (m)')
            ax.set_ylabel('Y (m)')
            ax.set_title(f'Lidar BEV - Frame {frame_idx}\nTime: {timestamp:.3f}s')
            ax.set_aspect('equal')
            ax.grid(True, alpha=0.3)
            plt.colorbar(scatter, ax=ax, label='Intensity')
            
            plt.savefig(os.path.join(self.dirs['lidar_img'], f'frame_{frame_idx:05d}.png'),
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
        except Exception:
            self.image_errors['lidar_img'] = self.image_errors.get('lidar_img', 0) + 1

    # ==========================================
    # 오디오 처리
    # ==========================================
    def _save_audio_spectrum(self, msg, frame_idx, timestamp):
        spectrum = np.array(msg.data)
        num_angles = len(spectrum)
        azimuth_grid = np.linspace(-180, 180, num_angles, endpoint=False)
        
        max_idx = np.argmax(spectrum)
        max_angle = azimuth_grid[max_idx]
        
        # 프레임별 CSV
        csv_path = os.path.join(self.dirs['acoustic_csv'], f'frame_{frame_idx:04d}.csv')
        df = pd.DataFrame({'angle': azimuth_grid, 'spectrum': spectrum})
        df.insert(0, 'max_angle', max_angle)
        df.insert(0, 'timestamp', timestamp)
        df.insert(0, 'frame', frame_idx)
        df.to_csv(csv_path, index=False)
        
        # Polar plot
        self._save_audio_plot(spectrum, azimuth_grid, frame_idx, max_angle, timestamp)

    def _save_audio_plot(self, spectrum, azimuth_grid, frame_idx, max_angle, timestamp):
        try:
            fig = plt.figure(figsize=(8, 6))
            ax = fig.add_subplot(111, projection='polar')
            ax.set_theta_zero_location('N')
            ax.set_theta_direction(-1)
            
            mask = (azimuth_grid >= AUDIO_ANGLE_MIN) & (azimuth_grid <= AUDIO_ANGLE_MAX)
            angles_filtered = azimuth_grid[mask]
            spectrum_filtered = spectrum[mask]
            
            ax.plot(np.deg2rad(angles_filtered), spectrum_filtered, color='red', linewidth=1.5)
            ax.fill(np.deg2rad(angles_filtered), spectrum_filtered, color='red', alpha=0.3)
            
            if AUDIO_ANGLE_MIN <= max_angle <= AUDIO_ANGLE_MAX:
                max_val = spectrum[np.argmax(spectrum)]
                ax.plot(np.deg2rad(max_angle), max_val, 'bo', markersize=10)
            
            ax.set_thetamin(AUDIO_ANGLE_MIN)
            ax.set_thetamax(AUDIO_ANGLE_MAX)
            ax.set_title(f'Audio - Frame {frame_idx} | Max: {max_angle:.1f}°\nTime: {timestamp:.3f}s')
            
            plt.savefig(os.path.join(self.dirs['acoustic_plots'], f'frame_{frame_idx:04d}.png'),
                       dpi=100, bbox_inches='tight')
            plt.close(fig)
        except Exception:
            self.image_errors['audio'] = self.image_errors.get('audio', 0) + 1

    def save_audio_wav(self):
        """수집된 audio/raw 데이터를 WAV 파일로 저장"""
        if not self.audio_raw_chunks:
            print("   ℹ️  오디오 raw 데이터 없음 (audio/raw 토픽 미녹화)")
            return
        
        print(f"\n🎵 오디오 WAV 저장 중...")
        
        # 전체 raw 데이터 결합
        all_audio = []
        for timestamp, raw_data in sorted(self.audio_raw_chunks, key=lambda x: x[0]):
            num_samples = len(raw_data) // AUDIO_CHANNELS
            if num_samples > 0:
                chunk = raw_data[:num_samples * AUDIO_CHANNELS].reshape(num_samples, AUDIO_CHANNELS)
                all_audio.append(chunk)
        
        if not all_audio:
            return
        
        audio_data = np.concatenate(all_audio, axis=0)
        
        # float32 → int16 변환 (WAV 저장용)
        audio_int16 = np.clip(audio_data * 32767, -32768, 32767).astype(np.int16)
        
        # 전체 WAV 파일 (16채널)
        wav_path = os.path.join(self.dirs['acoustic_wav'], 'audio_all_channels.wav')
        with wave.open(wav_path, 'w') as wf:
            wf.setnchannels(AUDIO_CHANNELS)
            wf.setsampwidth(2)  # 16-bit
            wf.setframerate(AUDIO_SAMPLE_RATE)
            wf.writeframes(audio_int16.tobytes())
        
        duration = len(audio_data) / AUDIO_SAMPLE_RATE
        print(f"   ✅ audio_all_channels.wav ({AUDIO_CHANNELS}ch, {duration:.1f}s, {len(audio_data)} samples)")
        
        # 채널별 개별 WAV (분석 편의용)
        for ch in range(AUDIO_CHANNELS):
            ch_path = os.path.join(self.dirs['acoustic_wav'], f'channel_{ch:02d}.wav')
            ch_data = audio_int16[:, ch]
            with wave.open(ch_path, 'w') as wf:
                wf.setnchannels(1)
                wf.setsampwidth(2)
                wf.setframerate(AUDIO_SAMPLE_RATE)
                wf.writeframes(ch_data.tobytes())
        
        print(f"   ✅ channel_00~{AUDIO_CHANNELS-1:02d}.wav (개별 채널)")

    # ==========================================
    # 일괄 CSV 저장
    # ==========================================
    def _save_bulk_csvs(self, topic_timestamps):
        print("\n💾 일괄 CSV 저장 중...")
        
        # 레이더 range/noise profile
        for radar_id in sorted(self.radar_ids):
            if radar_id == 'merged':
                continue
            
            rp_topic = f'/{radar_id}/range_profile'
            if rp_topic in self.raw_messages and self.raw_messages[rp_topic]:
                rows = []
                for i, tm in enumerate(self.raw_messages[rp_topic], 1):
                    rows.append([i, tm.timestamp] + list(tm.msg.data))
                if rows:
                    num_bins = len(rows[0]) - 2
                    columns = ['frame', 'timestamp'] + [f'bin_{j}' for j in range(num_bins)]
                    df = pd.DataFrame(rows, columns=columns)
                    save_path = os.path.join(self.dirs['radar'], radar_id, 'range_profile.csv')
                    os.makedirs(os.path.dirname(save_path), exist_ok=True)
                    df.to_csv(save_path, index=False)
                    print(f"   ✅ {radar_id}/range_profile.csv ({len(df)} frames)")
            
            np_topic = f'/{radar_id}/noise_profile'
            if np_topic in self.raw_messages and self.raw_messages[np_topic]:
                rows = []
                for i, tm in enumerate(self.raw_messages[np_topic], 1):
                    rows.append([i, tm.timestamp] + list(tm.msg.data))
                if rows:
                    num_bins = len(rows[0]) - 2
                    columns = ['frame', 'timestamp'] + [f'bin_{j}' for j in range(num_bins)]
                    df = pd.DataFrame(rows, columns=columns)
                    save_path = os.path.join(self.dirs['radar'], radar_id, 'noise_profile.csv')
                    os.makedirs(os.path.dirname(save_path), exist_ok=True)
                    df.to_csv(save_path, index=False)
                    print(f"   ✅ {radar_id}/noise_profile.csv ({len(df)} frames)")
        
        # 라이다 IMU
        imu_topic = '/ouster/imu'
        if imu_topic in self.raw_messages and self.raw_messages[imu_topic]:
            rows = []
            for i, tm in enumerate(self.raw_messages[imu_topic], 1):
                m = tm.msg
                rows.append([
                    i, tm.timestamp,
                    m.orientation.x, m.orientation.y, m.orientation.z, m.orientation.w,
                    m.angular_velocity.x, m.angular_velocity.y, m.angular_velocity.z,
                    m.linear_acceleration.x, m.linear_acceleration.y, m.linear_acceleration.z
                ])
            columns = ['frame', 'timestamp', 'ori_x', 'ori_y', 'ori_z', 'ori_w',
                       'ang_vel_x', 'ang_vel_y', 'ang_vel_z',
                       'lin_acc_x', 'lin_acc_y', 'lin_acc_z']
            df = pd.DataFrame(rows, columns=columns)
            df.to_csv(os.path.join(self.dirs['lidar'], 'imu_data.csv'), index=False)
            print(f"   ✅ lidar/imu_data.csv ({len(df)} frames)")

    def _save_sync_quality(self, rows):
        """동기화 품질 xlsx 저장"""
        if not rows:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            # openpyxl 없으면 CSV로 대체
            df = pd.DataFrame(rows)
            df.to_csv(os.path.join(self.output_dir, 'sync_quality.csv'), index=False)
            print(f"   ✅ sync_quality.csv ({len(rows)} frames)")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'sync_quality'
        
        # 스타일
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='D9E1F2')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center')
        warn_fill = PatternFill('solid', fgColor='FFFF00')
        
        # 헤더 (동적)
        if rows:
            headers = list(rows[0].keys())
        else:
            headers = ['frame', 'sync_time']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 데이터
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row_data.get(header, '')
                if isinstance(val, float):
                    val = round(val, 3)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align
                
                # diff_ms > 50ms 경고
                if 'diff_ms' in header and isinstance(val, (int, float)) and val > 50:
                    cell.fill = warn_fill
        
        # 열 너비
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(str(header)) + 2, 12)
        
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'
        
        save_path = os.path.join(self.output_dir, 'sync_quality.xlsx')
        wb.save(save_path)
        print(f"   ✅ sync_quality.xlsx ({len(rows)} frames)")

    def print_summary(self):
        print("\n" + "=" * 60)
        print("                    처리 완료 요약")
        print("=" * 60)
        print(f"  📁 출력 경로: {self.output_dir}")
        print(f"  🔄 동기화: {self.sync_rate_hz}Hz (후처리 nearest-neighbor)")
        print()
        
        for key, count in sorted(self.sync_stats.items()):
            print(f"  📊 {key}: {count} frames")
        
        if self.audio_raw_chunks:
            total_samples = sum(len(c[1]) // AUDIO_CHANNELS for c in self.audio_raw_chunks)
            print(f"  🎵 오디오 WAV: {total_samples} samples ({total_samples/AUDIO_SAMPLE_RATE:.1f}s)")
        
        print()
        total_errors = sum(self.image_errors.values())
        if total_errors > 0:
            print("  ⚠️ 이미지 생성 오류:")
            for key, count in self.image_errors.items():
                if count > 0:
                    print(f"     - {key}: {count}개")
        else:
            print("  ✅ 모든 파일 정상 생성")
        
        print("=" * 60)


# ==========================================
# 메인 실행부
# ==========================================
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='NLOS Multi-Sensor Rosbag Processor v3 (Post-Sync)')
    parser.add_argument('bag_path', nargs='?', default=None,
                       help='rosbag 폴더 경로 (없으면 최신 자동 탐색)')
    parser.add_argument('--sync-rate', type=float, default=10.0,
                       help='동기화 주파수 Hz (기본: 10.0)')
    
    args = parser.parse_args()
    
    if args.bag_path:
        bag_path = os.path.expanduser(args.bag_path)
    else:
        bag_path = get_latest_bag_path(ROSBAG_DIR)
    
    if bag_path and os.path.exists(bag_path):
        processor = BagProcessor(bag_path, sync_rate_hz=args.sync_rate)
        processor.process()
    else:
        print("❌ rosbag을 찾을 수 없습니다.")
        print(f"   검색 경로: {ROSBAG_DIR}")